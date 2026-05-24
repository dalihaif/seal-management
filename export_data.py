"""
印章档案目录 Excel → JSON 导入工具
用法: python export_data.py
说明: 从 Excel 文件读取印章目录数据，转换为 data.json 供系统导入
支持格式: .xlsx (推荐) / .xls (旧版)
"""
import openpyxl
import json
import os
import sys
from datetime import datetime, date, timedelta

# ==================== 配置区 ====================
EXCEL_PATH = r'E:/工作文档/档案室/08_印章档案/印章档案目录.xlsx'
# 如果上面路径的文件不存在，尝试以下备选路径：
BACKUP_PATHS = [
    r'E:/工作文档/档案室/08_印章档案/印章档案目录.xls',
    '印章档案目录.xlsx',
    '印章档案目录.xls',
]
OUTPUT_PATH = r'E:/工作文档/档案室/08_印章档案/印章管理系统/data.json'

def fmt_date(v):
    """格式化日期值"""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)):
        # Excel 序列号日期转换
        try:
            epoch = date(1899, 12, 30)
            return (epoch + timedelta(days=int(v))).strftime('%Y-%m-%d')
        except (OverflowError, ValueError):
            return str(v)
    return str(v).strip()


def load_xlsx(filepath):
    """使用 openpyxl 加载 xlsx 文件"""
    print(f"  [openpyxl] 正在加载: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_names = wb.sheetnames
    print(f"  [openpyxl] 工作表列表: {sheet_names}")

    # 优先使用第一个工作表
    ws = wb[sheet_names[0]]
    rows = list(ws.iter_rows(values_only=True))
    print(f"  [openpyxl] 读取到 {len(rows)} 行数据")
    return rows


def load_xls_old_format(filepath):
    """尝试加载旧版 .xls 格式（如果安装了 xlrd）"""
    try:
        import xlrd
        print(f"  [xlrd] 正在加载旧版Excel: {filepath}")
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        rows = []
        for row_idx in range(ws.nrows):
            row_values = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                value = cell.value
                # 处理日期类型
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(value, wb.datemode).strftime('%Y-%m-%d')
                    except Exception:
                        value = str(value)
                row_values.append(value)
            rows.append(tuple(row_values))
        print(f"  [xlrd] 读取到 {len(rows)} 行数据")
        return rows
    except ImportError:
        print("  [xlrd] 未安装 xlrd 库，无法读取 .xls 格式。请执行: pip install xlrd")
        return None
    except Exception as ex:
        print(f"  [xlrd] 读取失败: {ex}")
        return None


def main():
    print('=' * 60)
    print('  印章档案目录 Excel → JSON 导入工具')
    print('=' * 60)

    # ---- 1. 定位 Excel 文件 ----
    filepath = None
    if os.path.exists(EXCEL_PATH):
        filepath = EXCEL_PATH
    else:
        for bp in BACKUP_PATHS:
            if os.path.exists(bp):
                filepath = bp
                break

    if not filepath:
        print(f'\n❌ 错误: 找不到 Excel 文件!')
        print(f'   尝试的路径:')
        print(f'     - {EXCEL_PATH}')
        for bp in BACKUP_PATHS:
            print(f'     - {bp}')
        print(f'\n   请确认文件路径是否正确，或修改脚本中的 EXCEL_PATH 变量。')
        sys.exit(1)

    file_ext = os.path.splitext(filepath)[1].lower()
    file_size = os.path.getsize(filepath)
    print(f'\n📂 找到文件: {filepath}')
    print(f'   格式: {file_ext} | 大小: {file_size:,} 字节')

    # ---- 2. 根据格式选择加载方式 ----
    rows = None
    if file_ext == '.xlsx':
        try:
            rows = load_xlsx(filepath)
        except Exception as ex:
            print(f'\n❌ openpyxl 加载失败: {ex}')
            sys.exit(1)
    elif file_ext == '.xls':
        rows = load_xls_old_format(filepath)
        if rows is None:
            sys.exit(1)
    else:
        # 未知扩展名，先尝试 openpyxl
        print(f'  ⚠️ 未知扩展名 {file_ext}，尝试用 openpyxl 加载...')
        try:
            rows = load_xlsx(filepath)
        except Exception:
            rows = load_xls_old_format(filepath)

    if rows is None or len(rows) == 0:
        print('\n❌ 未能读取到任何数据')
        sys.exit(1)

    # ---- 3. 调试输出表头 ----
    header_row = rows[0] if rows else ()
    print(f'\n📋 表头 ({len(header_row)} 列):')
    for i, h in enumerate(header_row):
        display_val = str(h) if h is not None else '(空)'
        print(f'   [{i}] {display_val}')

    # ---- 4. 解析数据 ----
    data = []
    skip_count = 0
    for row_idx, row in enumerate(rows[1:], start=2):  # 跳过表头
        if row[0] is None:
            skip_count += 1
            continue

        # 确保行有足够的列
        padded_row = tuple(list(row) + [None] * max(0, 7 - len(row)))

        编号, 名称, 材质, 形状, 启用时间, 废止时间, 备注 = padded_row[:7]

        record = {
            'id': int(编号) if 编号 is not None and str(编号).strip() else 0,
            'name': str(名称).strip() if 名称 is not None else '',
            'material': str(材质).strip() if 材质 is not None else '',
            'shape': str(形状).strip() if 形状 is not None else '',
            'startDate': fmt_date(启用时间),
            'endDate': fmt_date(废止时间),
            'remark': str(备注).strip() if 备注 is not None else '',
            'status': '已废止' if (废止时间 is not None and str(废止时间).strip()) else '在用'
        }

        if record['name']:  # 只保留有名称的记录
            data.append(record)

    print(f'\n✅ 解析完成: 共 {len(rows)-1} 行原始数据, 有效记录 {len(data)} 条, 跳过空白 {skip_count} 条')

    # ---- 5. 显示前几条预览（验证中文是否正确） ----
    print(f'\n📝 数据预览 (前5条):')
    for i, d in enumerate(data[:5]):
        print(f'   [{i+1}] id={d["id"]} name="{d["name"]}" material="{d["material"]}" status={d["status"]}')

    # ---- 6. 写入 JSON ----
    output_dir = os.path.dirname(OUTPUT_PATH)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    output_size = os.path.getsize(OUTPUT_PATH)
    print(f'\n💾 已写入: {OUTPUT_PATH}')
    print(f'   大小: {output_size:,} 字节 | 编码: UTF-8 (ensure_ascii=False)')
    print(f'\n🎉 导出完成! 共 {len(data)} 条印章记录')


if __name__ == '__main__':
    main()
